from flask import Flask, render_template, request, redirect, session, send_file
import json, os
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import load_workbook, Workbook

app = Flask(__name__)
app.secret_key = 'my_secret_key'
app.config['UPLOAD_FOLDER'] = 'uploads'

# 사용자 불러오기
def load_users():
    with open('users.json', 'r', encoding='utf-8') as f:
        return json.load(f)

@app.route('/')
def home():
    return redirect('/login')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        user_id = request.form['id']
        users = load_users()
        if user_id in users:
            session['user'] = {
                'id': user_id,
                'name': users[user_id]['name'],
                'role': users[user_id]['role']
            }
            return redirect('/dashboard')
        return render_template('login.html', error="학번이 틀렸습니다.")
    return render_template('login.html')

@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect('/login')
    return render_template('dashboard.html', user=session['user'])

@app.route('/upload', methods=['POST'])
def upload():
    if 'user' not in session:
        return redirect('/login')

    user = session['user']
    photo = request.files['photo']
    if photo:
        today = datetime.now().strftime("%Y%m%d")
        now_time = datetime.now().strftime("%H:%M:%S")
        filename = f"{user['id']}_{user['name']}_{today}.jpg"
        filename = secure_filename(filename)

        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        photo.save(save_path)

        # 인증 기록 엑셀
        record_path = '조식인증기록.xlsx'
        try:
            wb1 = load_workbook(record_path)
            ws1 = wb1.active
        except:
            wb1 = Workbook()
            ws1 = wb1.active
            ws1.append(['날짜', '학번', '이름', '파일명', '시간'])
        ws1.append([today, user['id'], user['name'], filename, now_time])
        wb1.save(record_path)

        # 명렬표 엑셀에 O 체크
        namelist_path = '조식명렬표.xlsx'
        try:
            wb2 = load_workbook(namelist_path)
            ws2 = wb2.active
        except:
            wb2 = Workbook()
            ws2 = wb2.active
            ws2.append(['학번', '이름'])

        date_col = None
        for col in range(3, ws2.max_column + 2):
            val = ws2.cell(row=1, column=col).value
            if val == today:
                date_col = col
                break
            if val is None:
                ws2.cell(row=1, column=col).value = today
                date_col = col
                break

        for row in range(2, ws2.max_row + 1):
            if str(ws2.cell(row=row, column=1).value) == user['id']:
                ws2.cell(row=row, column=date_col).value = 'O'
                break
        wb2.save(namelist_path)

        return render_template('dashboard.html', user=user, message="인증 완료!")
    return "사진 업로드 실패"

@app.route('/download')
def download_excel():
    if 'user' not in session:
        return redirect('/login')
    if session['user']['role'] != 'admin':
        return "관리자만 접근할 수 있습니다."
    return send_file('조식인증기록.xlsx', as_attachment=True)

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

if __name__ == '__main__':
    # 외부에서 접근 가능하도록 0.0.0.0 바인딩
    app.run(host='0.0.0.0', port=5000, debug=True)
